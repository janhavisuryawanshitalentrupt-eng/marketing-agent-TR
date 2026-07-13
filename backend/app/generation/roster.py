"""Parse an uploaded EMPLOYEE ROSTER (CSV / .xlsx) and turn it into a magazine `issue` spec.

The user uploads a spreadsheet with a Name column + metric columns (Submissions, Interviews, Offers, Starts,
Productivity, …). We: (1) fuzzy-detect the name / office / numeric-metric columns, (2) rank everyone by a
performance score (a chosen column, or a weighted composite that favours outcomes), (3) feature the top
performer as the COVER champion and the next N as SPOTLIGHTS, with their real stats + auto-written copy.

Photos are NOT resolved here — the endpoint matches each featured name to a Folders employee (owner-scoped).
Pure + defensive: bad cells are skipped, never crash the build.
"""
from __future__ import annotations

import csv
import hashlib
import io
import math
import re
from collections import defaultdict

_NAME_EXACT = {"name", "employeename", "fullname", "employee", "person", "teammember", "candidate"}
_OFFICE_HINTS = ("office", "location", "branch", "city", "base", "region", "site")
_HEADLINES = ["Spark of Brilliance", "Setting the Pace", "Leading the Charge", "Raising the Bar",
              "Champion of the Month", "Driven to Deliver"]
_FEATURE_CAP = 24


def _norm(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (h or "").lower())


def _num(v) -> float | None:
    """Best-effort numeric parse: '1,234', '95%', '4.0' -> float; text / inf / nan / overflow -> None.
    (float() accepts 'inf'/'nan'/'1e400', so a NON-FINITE result is treated as text, never crashes _fmt.)"""
    if v is None:
        return None
    s = str(v).strip().replace(",", "").replace("%", "")
    if not s:
        return None
    f = None
    try:
        f = float(s)
    except ValueError:
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if m:
            try:
                f = float(m.group())
            except ValueError:
                f = None
    return f if (f is not None and math.isfinite(f)) else None


def _numeric_frac(vals) -> float:
    non = [v for v in vals if str(v).strip()]
    return (sum(1 for v in non if _num(v) is not None) / len(non)) if non else 0.0


def _text_frac(vals) -> float:
    non = [v for v in vals if str(v).strip()]
    return (sum(1 for v in non if _num(v) is None) / len(non)) if non else 0.0


def _fmt(v) -> str:
    n = _num(v)
    s = str(v).strip()
    if n is None or not math.isfinite(n):   # non-numeric / inf / nan -> show the raw text (never int(inf))
        return s[:12]
    # Thousands-separated, up to 2 decimals — NEVER f"{n:g}" (that is 6 significant digits: it drops
    # cents, invents round figures, and turns big sums into scientific notation e.g. 1234567.89->'1.23457e+06').
    if n == int(n):
        whole = f"{int(n):,}"
    else:
        whole = f"{round(n, 2):,.2f}"
    return whole + "%" if "%" in s else whole


# ---- parsing ------------------------------------------------------------------------------------
def parse_roster(filename: str, data: bytes) -> list[dict]:
    """Return a list of row dicts {header: cell}. Handles .xlsx (openpyxl) and CSV/TSV (stdlib)."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(data)
    return _parse_csv(data)


def _rows_to_dicts(header: list, rows: list[list]) -> list[dict]:
    # Coerce EVERY header cell to str first — an .xlsx header row can hold ints (a year like 2024), floats or
    # datetimes, and (h or "").strip() would then raise AttributeError on the non-string.
    header = [(str(h).strip() if h is not None else "") for h in header]
    # de-dupe blank/duplicate headers so keys stay unique
    seen: dict[str, int] = {}
    keys: list[str] = []
    for i, h in enumerate(header):
        key = h or f"col{i + 1}"
        if key in seen:
            seen[key] += 1
            key = f"{key} ({seen[key]})"
        else:
            seen[key] = 0
        keys.append(key)
    out = []
    for r in rows:
        out.append({keys[i]: (str(r[i]).strip() if i < len(r) and r[i] is not None else "")
                    for i in range(len(keys))})
    return out


def _parse_csv(data: bytes) -> list[dict]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        text = data.decode("utf-8", "replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return []
    return _rows_to_dicts(rows[0], rows[1:])


def _parse_xlsx(data: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = [r for r in ws.iter_rows(values_only=True)
                if r is not None and any(c is not None and str(c).strip() for c in r)]
    finally:
        wb.close()
    if not rows:
        return []
    return _rows_to_dicts(list(rows[0]), [list(r) for r in rows[1:]])


# ---- column detection + scoring -----------------------------------------------------------------
def _detect_columns(rows: list[dict]):
    headers = list(rows[0].keys()) if rows else []
    name_col = next((h for h in headers if _norm(h) in _NAME_EXACT), None)
    if not name_col:
        name_col = next((h for h in headers if "name" in _norm(h)), None)
    if not name_col:  # fallback: the first mostly-text column
        name_col = next((h for h in headers if _text_frac([r.get(h, "") for r in rows]) > 0.6), None)
    office_col = next((h for h in headers
                       if h != name_col and any(k in _norm(h) for k in _OFFICE_HINTS)), None)
    metric_cols = [h for h in headers
                   if h and h not in (name_col, office_col)
                   and _numeric_frac([r.get(h, "") for r in rows]) >= 0.5]
    return name_col, office_col, metric_cols


def _weight(h: str) -> float:
    n = _norm(h)
    if any(k in n for k in ("start", "join", "placement", "onboard", "hire")):
        return 5.0
    if "offer" in n:
        return 4.0
    if "interview" in n:
        return 1.5
    if any(k in n for k in ("productivity", "score", "rating", "points", "efficiency")):
        return 2.0
    if any(k in n for k in ("submission", "sub", "cv", "profile", "sourced")):
        return 0.3
    return 1.0


def _score(row: dict, metric_cols: list[str], rank_by: str) -> float:
    if rank_by:
        for h in metric_cols:
            if _norm(h) == _norm(rank_by):
                v = _num(row.get(h))
                return v if v is not None else -1.0
    total = 0.0
    for h in metric_cols:
        v = _num(row.get(h))
        if v is not None:
            total += v * _weight(h)
    return total


# ---- copy templates -----------------------------------------------------------------------------
def _champion_copy(first: str, stats: list[dict]) -> tuple[str, str]:
    hl = _HEADLINES[int(hashlib.md5((first or "x").encode("utf-8")).hexdigest()[:6], 16) % len(_HEADLINES)]
    if len(stats) >= 2:
        tg = (f"With {stats[0]['value']} {stats[0]['label'].lower()} and {stats[1]['value']} "
              f"{stats[1]['label'].lower()}, {first} set a new benchmark this month.")
    elif stats:
        tg = f"With {stats[0]['value']} {stats[0]['label'].lower()}, {first} led the way this month."
    else:
        tg = f"{first} delivered standout results this month."
    return hl, tg


def _spotlight_blurb(first: str, stats: list[dict]) -> str:
    if stats:
        top = ", ".join(f"{s['value']} {s['label'].lower()}" for s in stats[:2])
        return (f"{first} continues to set the benchmark — {top} this month reflect real consistency and "
                "quality delivery.")
    return f"{first} delivered consistent, quality results — a standout month for the team."


# ---- public: roster -> issue --------------------------------------------------------------------
def build_issue(rows: list[dict], *, theme: str = "", title: str = "Talentrupt Times", edition: str = "",
                editorial: str = "", feature_count: int = 0, rank_by: str = ""):
    """Analyse the roster and return (issue_dict_without_photos, featured_names, columns_info).

    Raises ValueError('no_name_column' | 'no_rows') for a caller-friendly 400.
    """
    rows = [r for r in rows if any((str(v) or "").strip() for v in r.values())]
    if not rows:
        raise ValueError("no_rows")
    name_col, office_col, metric_cols = _detect_columns(rows)
    if not name_col:
        raise ValueError("no_name_column")
    people = [r for r in rows if (r.get(name_col) or "").strip()]
    if not people:
        raise ValueError("no_rows")

    ranked = sorted(people, key=lambda r: _score(r, metric_cols, rank_by), reverse=True)
    limit = min(feature_count, _FEATURE_CAP) if feature_count and feature_count > 0 else _FEATURE_CAP
    ranked = ranked[:limit]

    def stats_for(r: dict, n: int) -> list[dict]:
        out = []
        for h in metric_cols:
            if len(out) >= n:
                break
            v = r.get(h, "")
            if str(v).strip():
                out.append({"label": h, "value": _fmt(v)})
        return out

    champ = ranked[0]
    cname = (champ.get(name_col) or "").strip()
    cfirst = cname.split()[0] if cname else "Our champion"
    cstats = stats_for(champ, 6)
    headline, tagline = _champion_copy(cfirst, cstats)

    issue: dict = {
        "title": (title or "Talentrupt Times").strip(),
        "edition": edition.strip(),
        "theme": theme.strip(),
        "editorial": editorial.strip(),
        "cover": {"name": cname, "role": "", "headline": headline, "tagline": tagline, "stats": cstats},
        "spotlights": [],
    }
    for r in ranked[1:]:
        nm = (r.get(name_col) or "").strip()
        if not nm:
            continue
        sstats = stats_for(r, 4)
        issue["spotlights"].append({
            "name": nm, "role": "",
            "office": (r.get(office_col) or "").strip() if office_col else "",
            "blurb": _spotlight_blurb(nm.split()[0], sstats), "stats": sstats,
        })

    featured = [cname] + [s["name"] for s in issue["spotlights"]]
    columns = {"name": name_col, "office": office_col, "metrics": metric_cols}
    return issue, featured, columns


# =================================================================================================
# AWARD-REPORT format  (a real HR "FNL report" workbook: an awards leaderboard tab laid out in
# side-by-side blocks — Margin Champions / Placements Powerhouse / Efficiency Star / Category
# Champions — PLUS a raw "Deal sheet" of one row per placement. We read the curated podiums AS-IS
# and enrich each featured person with stats aggregated from the raw deals.)
# =================================================================================================

# Award-block titles we recognise: (normalised-substring, display, unit).
_AWARD_BLOCKS = [
    ("marginchampion", "Margin Champions", "margin"),
    ("placementspowerhouse", "Placements Powerhouse", "placements"),
    ("placementpowerhouse", "Placements Powerhouse", "placements"),
    ("efficiencystar", "Efficiency Star", "efficiency"),
]
_CATEGORY_KEY = "categorychampion"
_SUBCATS = {"li": "LI", "nontech": "Non-Tech", "tech": "Tech"}
_PODIUM = 5          # winners parsed per block (renderer shows the top 3)
_MAX_AWARDS = 12     # hard cap on award blocks (defensive)
_DEFAULT_SPOTLIGHTS = 6
_SCAN = 16           # rows scanned for block titles / the deal header (tolerates banner/spacer rows)


def _value_display(unit: str, raw: str, n) -> str:
    """A clean big-number string for a podium winner ('$18.55', '92.85', '20')."""
    if n is None:
        return (raw or "").strip()[:12] or "—"
    return ("$" + _fmt(n)) if unit == "efficiency" else _fmt(n)


def _rank1_at(grid: list[list[str]], r: int, c: int) -> bool:
    """True if cell (r, c) holds the rank number 1 — the structural signature of a real award block's
    first data row directly under its title. Guards against stray award words in ordinary data cells."""
    return r < len(grid) and c < len(grid[r]) and _num(grid[r][c]) == 1


def parse_workbook(filename: str, data: bytes) -> dict[str, list[list[str]]]:
    """Return {sheet_name: grid_of_str_cells}. .xlsx -> every sheet; CSV/TSV -> one 'Sheet1' grid."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            out: dict[str, list[list[str]]] = {}
            for sn in wb.sheetnames:
                ws = wb[sn]
                out[str(sn)] = [
                    [("" if c is None else str(c).strip()) for c in row]
                    for row in ws.iter_rows(values_only=True)
                ]
            return out
        finally:
            wb.close()
    # CSV / TSV -> a single grid (reuse the sniffing/decoding from _parse_csv)
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        text = data.decode("utf-8", "replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    grid = [[(c or "").strip() for c in r] for r in reader]
    return {"Sheet1": grid}


def grid_to_rows(grid: list[list[str]]) -> list[dict]:
    """Turn a raw grid (header row + data rows) into row-dicts for the flat-table `build_issue`."""
    if not grid:
        return []
    return _rows_to_dicts(grid[0], grid[1:])


def best_flat_grid(sheets: dict[str, list[list[str]]]) -> list[list[str]]:
    """Pick the sheet most likely to be the data table — the one with the most non-blank rows — so a
    multi-sheet workbook whose first tab is a cover/blank sheet still finds the real roster."""
    best: list[list[str]] = []
    best_n = -1
    for grid in sheets.values():
        n = sum(1 for r in grid if any((c or "").strip() for c in r))
        if n > best_n:
            best, best_n = grid, n
    return best


def is_award_format(sheets: dict[str, list[list[str]]]) -> bool:
    """True only when a sheet parses into >=2 STRUCTURAL award blocks (a title with rank-1 data below).
    A plain roster that merely mentions an award word in a data cell parses 0 blocks -> False, so it
    correctly falls through to the flat-table builder instead of being garbled by the award builder."""
    grid = _find_award_grid(sheets)
    return bool(grid) and len(parse_award_sheet(grid)) >= 2


def _find_award_grid(sheets: dict[str, list[list[str]]]) -> list[list[str]] | None:
    best, best_hits = None, 0
    for grid in sheets.values():
        hits = 0
        for row in grid[:_SCAN]:
            for cell in row:
                nc = _norm(cell)
                if any(k in nc for k, _d, _u in _AWARD_BLOCKS) or _CATEGORY_KEY in nc:
                    hits += 1
        if hits > best_hits:
            best, best_hits = grid, hits
    return best


def _find_deal_sheet(sheets: dict[str, list[list[str]]]):
    """Return (grid, header_row_index) for the raw placements sheet, or None."""
    for grid in sheets.values():
        for ri, row in enumerate(grid[:_SCAN]):
            nn = [_norm(c) for c in row]
            if "recruiter" in nn and "spread" in nn:
                return grid, ri
    return None


def _read_triples(grid: list[list[str]], r0: int, c: int, unit: str, take: int = _PODIUM) -> list[dict]:
    """Read [rank, name, value] rows downward from (r0, c). Stops at the first blank name."""
    out: list[dict] = []
    r = r0
    while r < len(grid) and len(out) < take:
        row = grid[r]
        name = row[c + 1].strip() if c + 1 < len(row) else ""
        raw = row[c + 2].strip() if c + 2 < len(row) else ""
        if not name:
            break
        out.append({"rank": len(out) + 1, "name": name,
                    "value": _value_display(unit, raw, _num(raw)), "num": _num(raw), "photo": None})
        r += 1
    return out


def parse_award_sheet(grid: list[list[str]]) -> list[dict]:
    """Parse the block-layout awards tab into [{title, unit, winners:[{rank,name,value,...}]}].

    A block is only accepted when the cell directly below its title holds rank 1 (`_rank1_at`) — so a
    stray 'Efficiency Star' in some data cell can't masquerade as a block. Category sub-headers are read
    only in columns at/after the 'Category champions' title, so a real winner surnamed 'Li'/'Tech' in a
    left-hand block can't be mistaken for a category sub-header.
    """
    awards: list[dict] = []
    for ri, row in enumerate(grid[:_SCAN]):
        for ci, cell in enumerate(row):
            nc = _norm(cell)
            for key, disp, unit in _AWARD_BLOCKS:
                if key in nc and _rank1_at(grid, ri + 1, ci):
                    winners = _read_triples(grid, ri + 1, ci, unit)
                    if winners and not any(a["title"] == disp for a in awards):
                        awards.append({"title": disp, "unit": unit, "winners": winners})
            if _CATEGORY_KEY in nc and ri + 1 < len(grid):
                for cj in range(ci, len(grid[ri + 1])):
                    label = _SUBCATS.get(_norm(grid[ri + 1][cj]))
                    if label and _rank1_at(grid, ri + 2, cj):
                        winners = _read_triples(grid, ri + 2, cj, "category")
                        title = f"Category Champion — {label}"
                        if winners and not any(a["title"] == title for a in awards):
                            awards.append({"title": title, "unit": "category", "winners": winners})
            if len(awards) >= _MAX_AWARDS:
                return awards
    return awards


def aggregate_deals(grid: list[list[str]], header_row: int) -> dict[str, dict]:
    """Group the raw deal rows by Recruiter -> {norm_name: {name, placements, margin, efficiency}}."""
    header = [_norm(c) for c in grid[header_row]]
    R = header.index("recruiter") if "recruiter" in header else None
    S = header.index("spread") if "spread" in header else None
    if R is None:
        return {}
    placements: dict[str, int] = defaultdict(int)
    margin: dict[str, float] = defaultdict(float)
    display: dict[str, str] = {}
    for row in grid[header_row + 1:]:
        if R >= len(row):
            continue
        raw = row[R].strip()
        if not raw or raw.lower() == "na":
            continue
        key = _norm(raw)
        if not key:
            continue
        display.setdefault(key, raw)
        placements[key] += 1
        v = _num(row[S]) if (S is not None and S < len(row)) else None
        if v is not None:
            margin[key] += v
    out: dict[str, dict] = {}
    for key, p in placements.items():
        m = round(margin[key], 2)
        out[key] = {"name": display[key], "placements": p, "margin": m,
                    "efficiency": round(m / p, 2) if p else 0.0}
    return out


def build_award_issue(sheets: dict[str, list[list[str]]], *, theme: str = "",
                      title: str = "Talentrupt Times", edition: str = "", editorial: str = "",
                      feature_count: int = 0):
    """Build a magazine `issue` from an award-report workbook.

    Returns (issue, featured_names, meta). `issue` carries an extra `awards` list of podium blocks
    (each winner gets a `photo` slot the endpoint fills) plus per-person `spotlights` (stat pages).
    Raises ValueError('no_awards') if the awards tab can't be read.
    """
    award_grid = _find_award_grid(sheets)
    awards = parse_award_sheet(award_grid) if award_grid else []
    if not awards:
        raise ValueError("no_awards")

    deal = _find_deal_sheet(sheets)
    agg = aggregate_deals(*deal) if deal else {}

    def stats_for(name: str) -> list[dict]:
        a = agg.get(_norm(name))
        if not a:
            return []
        return [
            {"label": "Placements", "value": _fmt(a["placements"])},
            {"label": "Total Margin", "value": _fmt(a["margin"])},
            {"label": "Avg / Placement", "value": "$" + _fmt(a["efficiency"])},
        ]

    # Cover champion = the Margin Champions #1 (or the first award's #1 if margin is absent).
    lead = next((a for a in awards if a["unit"] == "margin"), awards[0])
    champ = lead["winners"][0]
    cname = champ["name"]
    cfirst = cname.split()[0] if cname else "Our champion"
    cstats = stats_for(cname) or [{"label": lead["title"], "value": champ["value"]}]
    headline, tagline = _champion_copy(cfirst, cstats)

    issue: dict = {
        "title": (title or "Talentrupt Times").strip(),
        "edition": edition.strip(),
        "theme": theme.strip(),
        "editorial": editorial.strip(),
        "cover": {"name": cname, "role": "", "headline": headline, "tagline": tagline, "stats": cstats},
        "awards": awards,
        "spotlights": [],
    }

    # Per-person STAT PAGES: distinct winners (excluding the cover champion), richest first.
    seen = {_norm(cname)}
    distinct: list[str] = []
    for a in awards:
        for w in a["winners"]:
            k = _norm(w["name"])
            if k and k not in seen:
                seen.add(k)
                distinct.append(w["name"])
    distinct.sort(key=lambda n: agg.get(_norm(n), {}).get("margin", 0.0), reverse=True)
    cap = feature_count if (feature_count and feature_count > 0) else _DEFAULT_SPOTLIGHTS
    for nm in distinct[:min(cap, _FEATURE_CAP)]:
        issue["spotlights"].append({
            "name": nm, "role": "", "office": "",
            "blurb": _spotlight_blurb(nm.split()[0] if nm else "", stats_for(nm)),
            "stats": stats_for(nm),
        })

    # Every podium winner (union) — the endpoint needs to attach a photo to each.
    all_featured: list[str] = []
    seen2: set[str] = set()
    for a in awards:
        for w in a["winners"][:3]:
            k = _norm(w["name"])
            if k and k not in seen2:
                seen2.add(k)
                all_featured.append(w["name"])
    featured = [cname] + [s["name"] for s in issue["spotlights"]]
    meta = {
        "format": "award",
        "awards": [{"title": a["title"], "winners": [w["name"] for w in a["winners"][:3]]} for a in awards],
        "sheets": list(sheets.keys()),
        "deal_sheet": bool(deal),
        "all_featured": all_featured,
    }
    return issue, featured, meta
